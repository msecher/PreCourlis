import os

from openpyxl import load_workbook

from test import DATA_PATH


def test_visu_profils():
    path = os.path.join(DATA_PATH, "input", "Exemple_VisuProfils_v2.xlsm")
    wb = load_workbook(path)
    ws = wb.get_sheet_by_name("Données")

    len([cell for cell in ws["A"] if cell.value])

    return
